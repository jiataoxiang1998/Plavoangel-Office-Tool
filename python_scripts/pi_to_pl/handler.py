import pandas as pd
import os
import io
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Color, Border, Side


class PItoPLHandler:
    def __init__(self):
        self.sales_contracts = {}
        self.selected_articles = set()

    def add_contract(self, file_path):
        """添加销售合同，提取货号和详细数据"""
        try:
            parsed_data = self.read_sales_contract(file_path)
            article_numbers = []
            seen = set()
            for item in parsed_data:
                art = item.get('article')
                if art and art not in seen:
                    seen.add(art)
                    article_numbers.append(art)
            
            self.sales_contracts[file_path] = {
                'article_numbers': article_numbers,
                'data': parsed_data
            }
            
            for art in article_numbers:
                self.selected_articles.add(art)
            
            return True, article_numbers
        except Exception as e:
            return False, str(e)

    def remove_contract(self, file_path):
        """移除销售合同"""
        if file_path in self.sales_contracts:
            removed_articles = self.sales_contracts[file_path]['article_numbers']
            del self.sales_contracts[file_path]
            self.selected_articles = self.selected_articles - set(removed_articles)
            return True
        return False

    def get_all_article_numbers(self):
        """获取所有货号（保持销售合同中的顺序）"""
        all_articles = []
        seen = set()
        for contract in self.sales_contracts.values():
            for art in contract['article_numbers']:
                if art not in seen:
                    seen.add(art)
                    all_articles.append(art)
        return all_articles

    def toggle_article_selection(self, article_number):
        """切换货号选中状态"""
        if article_number in self.selected_articles:
            self.selected_articles.discard(article_number)
            return False
        else:
            self.selected_articles.add(article_number)
            return True

    def remove_article_selection(self, article_number):
        """移除货号选中状态"""
        self.selected_articles.discard(article_number)

    def get_selected_articles(self):
        """获取选中的货号"""
        return list(self.selected_articles)

    def is_article_selected(self, article_number):
        """检查货号是否被选中"""
        return article_number in self.selected_articles

    def get_selected_articles_in_order(self):
        """获取按销售合同顺序排列的选中货号"""
        all_articles = self.get_all_article_numbers()
        return [art for art in all_articles if art in self.selected_articles]

    def generate_packing_list(self, output_path, selected_articles=None):
        """生成装箱单"""
        if selected_articles is None:
            selected_articles = self.get_selected_articles_in_order()
        
        if not selected_articles:
            raise Exception("请先选择货号")
        
        all_data = []
        selected_set = set(str(a) for a in selected_articles)
        
        for contract_path, contract_data in self.sales_contracts.items():
            parsed_data = contract_data['data']
            images = self.get_images_from_excel(contract_path, column=6)
            matching_rows = []
            
            if isinstance(parsed_data, list):
                for item in parsed_data:
                    if item.get('article') in selected_set:
                        matching_rows.append(item)
            
            if matching_rows:
                all_data.append({
                    'file_path': contract_path,
                    'rows': matching_rows,
                    'images': images
                })
        
        if not all_data:
            raise Exception("未找到选中的货号数据")
        
        base_name, ext = os.path.splitext(output_path)
        if not ext:
            ext = '.xlsx'
            output_path = base_name + ext
        
        wb = Workbook()
        ws = wb.active
        ws.title = "PL"
        
        # 冻结窗格在第5行
        ws.freeze_panes = 'A5'
        
        # 设置视图为分页预览（显示灰色周边）
        ws.sheet_view.view = 'pageBreakPreview'
        
        # 隐藏网格线
        ws.sheet_view.showGridLines = False
        
        # Row 1: 标题
        ws.merge_cells('I1:I1')
        ws.column_dimensions['I'].width = 15
        ws.row_dimensions[1].height = 39
        cell_i1 = ws['I1']
        cell_i1.value = '步步新出货箱单'
        cell_i1.font = Font(name='Trebuchet MS', bold=True, size=20)
        cell_i1.alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 2: 出货信息
        # Row 2 合并 A2:B2
        ws.merge_cells('A2:B2')
        ws['A2'] = '出货日:'
        ws['C2'] = 45960
        ws['E2'] = '拼箱进仓编号：'
        ws['H2'] = '步步新'
        
        for col in ['A2', 'C2', 'E2', 'H2']:
            ws[col].font = Font(name='Trebuchet MS', bold=True, size=18)
        
        ws['A2'].alignment = Alignment(horizontal='left')
        ws['C2'].alignment = Alignment(horizontal='left')
        ws['H2'].alignment = Alignment(horizontal='center')
        
        # Row 3-4: 列标题 (按模板格式，包含材质列D)
        ws.merge_cells('A3:A4')
        ws.merge_cells('B3:B4')
        ws.merge_cells('C3:C4')
        ws.merge_cells('D3:D4')  # 材质
        ws.merge_cells('E3:E4')  # 尺寸
        ws.merge_cells('F3:H4')  # 图片
        ws.merge_cells('I3:I4')
        ws.merge_cells('J3:J4')
        ws.merge_cells('K3:K4')
        ws.merge_cells('L3:L4')
        ws.merge_cells('M3:M4')
        ws.merge_cells('N3:N4')
        ws.merge_cells('O3:O4')
        ws.merge_cells('P3:R3')  # MEAS (L,W,H)
        ws.merge_cells('S3:S4')
        ws.merge_cells('T3:T4')
        ws.merge_cells('U3:U4')
        ws.merge_cells('V3:V4')
        ws.merge_cells('W3:W4')
        ws.merge_cells('X3:X4')
        
        # Row 3: 列标题 (包含材质列D)
        header_font = Font(name='Trebuchet MS', bold=True, size=10, color=Color(indexed=9))
        header_fill = PatternFill(patternType='solid', fgColor=Color(theme=3, tint=-0.499984740745262))
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 所有需要背景色的列（Row 3-4合并的列）
        header_cols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24]
        for col in header_cols:
            cell = ws.cell(row=3, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        headers_row3 = {
            1: 'No.', 2: 'ITEM #', 3: 'DESCRIPTION', 4: '材质', 5: '尺寸',
            9: 'COLOR', 10: '每箱颜色比例(订单要求)', 11: '清装', 12: '尾箱',
            13: '件数', 14: 'PCS每箱数量', 15: 'QTY/SET',
            16: 'MEAS. (M)', 19: 'CBM/总体积', 20: '单位净重', 21: '单位毛重',
            22: '总净重', 23: '总毛重', 24: '备注'
        }
        for col, header in headers_row3.items():
            ws.cell(row=3, column=col, value=header)
        
        # Row 4: 所有合并单元格的列都需要背景色和边框
        for col in header_cols:
            cell = ws.cell(row=4, column=col)
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Row 4: L, W, H
        ws['P4'] = 'L'
        ws['Q4'] = 'W'
        ws['R4'] = 'H'
        lwh_font = Font(name='Trebuchet MS', bold=True, size=10, color=Color(indexed=9))
        lwh_alignment = Alignment(horizontal='center', vertical='center')
        for col in ['P4', 'Q4', 'R4']:
            ws[col].font = lwh_font
            ws[col].alignment = lwh_alignment
            ws[col].border = thin_border
        
        # 数据行
        row_no = 1
        current_row = 5
        
        # 按货号分组
        articles_data = {}
        for contract_info in all_data:
            images = contract_info.get('images', {})
            for item in contract_info['rows']:
                article = item.get('article')
                if article not in selected_set:
                    continue
                if article not in articles_data:
                    articles_data[article] = {'items': [], 'images': images}
                articles_data[article]['items'].append(item)
        
        for article in selected_articles:
            if article not in articles_data:
                continue
            
            items = articles_data[article]['items']
            images = articles_data[article]['images']
            
            # 找到主货号行（row_type == 'main'）
            main_item = None
            color_items = []
            for item in items:
                if item.get('row_type') == 'main':
                    main_item = item
                else:
                    color_items.append(item)
            
            if not main_item:
                continue
            
            # 获取数据
            box_count = main_item.get('box_count', 0) or 0  # 总箱数 -> 件数
            box_pcs = main_item.get('box_pcs', 0) or 0  # 装箱量 -> PCS每箱数量
            quantity = main_item.get('quantity', 0) or 0  # 总数量 -> QTY/SET
            
            # 计算总配比：主行的R值 + 所有颜色明细行的R值
            total_ratio = (main_item.get('color_ratio', 0) or 0)
            for ci in color_items:
                total_ratio += (ci.get('color_ratio', 0) or 0)
            
            # 数据行样式
            data_font = Font(name='Trebuchet MS', size=11)
            data_alignment_center = Alignment(horizontal='center', vertical='center')
            data_alignment_left = Alignment(horizontal='left', vertical='center')
            data_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 主货号行
            ws.cell(row=current_row, column=1, value=row_no).font = data_font
            ws.cell(row=current_row, column=1).alignment = data_alignment_center
            ws.cell(row=current_row, column=2, value=article).font = data_font
            ws.cell(row=current_row, column=2).alignment = data_alignment_center
            ws.cell(row=current_row, column=3, value=str(main_item.get('description', '')) if main_item.get('description') else 'PU手机包+肩带').font = data_font
            ws.cell(row=current_row, column=3).alignment = data_alignment_center
            ws.cell(row=current_row, column=4, value='PU').font = data_font
            ws.cell(row=current_row, column=4).alignment = data_alignment_center
            ws.cell(row=current_row, column=5, value=str(main_item.get('size', '')) if main_item.get('size') else '').font = data_font
            ws.cell(row=current_row, column=5).alignment = data_alignment_center
            
            color_code = str(main_item.get('color_code', '')) if main_item.get('color_code') else ''
            main_ratio = main_item.get('color_ratio', 0) or 0
            
            # 每箱颜色比例 = (当前颜色配比 / 总配比) * 装箱量
            if total_ratio > 0 and box_pcs > 0:
                main_color_ratio = round(main_ratio / total_ratio * box_pcs)
            else:
                main_color_ratio = main_ratio
            
            ws.cell(row=current_row, column=9, value=color_code).font = data_font
            ws.cell(row=current_row, column=9).alignment = data_alignment_left
            ws.cell(row=current_row, column=10, value=main_color_ratio).font = data_font
            ws.cell(row=current_row, column=10).alignment = data_alignment_center
            
            # 件数 = 总箱数, PCS每箱数量 = 装箱量, QTY/SET = 件数*PCS每箱数量
            ws.cell(row=current_row, column=13, value=box_count).font = data_font
            ws.cell(row=current_row, column=13).alignment = data_alignment_center
            ws.cell(row=current_row, column=14, value=box_pcs).font = data_font
            ws.cell(row=current_row, column=14).alignment = data_alignment_center
            ws.cell(row=current_row, column=15, value=f'=M{current_row}*N{current_row}').font = data_font
            ws.cell(row=current_row, column=15).alignment = data_alignment_center
            
            # MEAS L/W/H, CBM, 单位净重, 单位毛重, 总净重, 总毛重 - 销售合同中没有，保留为空
            
            # 添加图片
            excel_row = main_item.get('original_row', 0) + 1
            if excel_row in images:
                img_data = images[excel_row]
                if img_data:
                    buffer = io.BytesIO(img_data)
                    img = XLImage(buffer)
                    img.width = 180
                    img.height = 180
                    ws.add_image(img, f'F{current_row}')
            
            row_no += 1
            
            # 颜色明细行
            for color_item in color_items:
                current_row += 1
                c_code = str(color_item.get('color_code', '')) if color_item.get('color_code') else ''
                c_ratio = color_item.get('color_ratio', 0) or 0
                
                # 每箱颜色比例 = (当前颜色配比 / 总配比) * 装箱量
                if total_ratio > 0 and box_pcs > 0:
                    color_ratio_calc = round(c_ratio / total_ratio * box_pcs)
                else:
                    color_ratio_calc = c_ratio
                
                ws.cell(row=current_row, column=9, value=c_code).font = data_font
                ws.cell(row=current_row, column=9).alignment = data_alignment_left
                ws.cell(row=current_row, column=10, value=color_ratio_calc).font = data_font
                ws.cell(row=current_row, column=10).alignment = data_alignment_center
            
            # 货号行合并（从主行到最后一个颜色行）
            total_rows_for_article = 1 + len(color_items)  # 主行 + 颜色明细行
            start_row = current_row - len(color_items)  # 主行所在行
            end_row = current_row
            
            # 为数据行添加边框
            for row in range(start_row, end_row + 1):
                for col in range(1, 25):
                    ws.cell(row=row, column=col).border = data_border
            
            # 需要合并的列（F,G,H需要合并为一个单元格）
            merge_cols_single = ['A', 'B', 'C', 'D', 'E', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
            for col_letter in merge_cols_single:
                ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
            
            # FGH三列合并为一个单元格
            ws.merge_cells(f'F{start_row}:H{end_row}')
            
            current_row += 1
        
        # TOTAL 行
        # 获取数据起始行（应该是第5行）
        data_start_row = 5
        total_row = current_row
        
        # 合并 A:H
        ws.merge_cells(f'A{total_row}:H{total_row}')
        ws[f'A{total_row}'] = 'TOTAL:'
        
        total_font = Font(name='Trebuchet MS', bold=True, size=10)
        total_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFFFFF00'))
        total_alignment_right = Alignment(horizontal='right', vertical='center')
        total_alignment_center = Alignment(horizontal='center', vertical='center')
        
        # A 列
        ws[f'A{total_row}'].font = total_font
        ws[f'A{total_row}'].fill = total_fill
        ws[f'A{total_row}'].alignment = total_alignment_right
        ws[f'A{total_row}'].border = thin_border
        
        # 设置其他列的公式
        ws[f'M{total_row}'] = f'=SUM(M{data_start_row}:M{total_row-1})'
        ws[f'M{total_row}'].font = total_font
        ws[f'M{total_row}'].fill = total_fill
        ws[f'M{total_row}'].alignment = total_alignment_center
        ws[f'M{total_row}'].border = thin_border
        
        ws[f'O{total_row}'] = f'=SUM(O{data_start_row}:O{total_row-1})'
        ws[f'O{total_row}'].font = total_font
        ws[f'O{total_row}'].fill = total_fill
        ws[f'O{total_row}'].alignment = total_alignment_center
        ws[f'O{total_row}'].border = thin_border
        
        # S(CBM), T(单位净重), U(单位毛重), V(总净重), W(总毛重) 销售合同中没有数据，保留为空
        
        current_row += 1
        
        # 备注行 (保持为空)
        current_row += 1
        
        # 创建对账单工作表
        self._create_duizhangdan_sheet(wb, all_data, selected_articles)
        
        wb.save(output_path)
        return output_path

    def read_sales_contract(self, file_path):
        """读取销售合同文件，区分主货号行和颜色明细行
        
        主货号行：Article# (C列索引2) 有值且以 'CCW' 开头
        颜色明细行：Article# 为空但 L列(索引11)有值，货号继承自上一个主行
        
        列索引:
        - A列(0): 序号
        - B列(1): 客户货号
        - C列(2): Article# (货号)
        - E列(4): 产品描述
        - J列(9): 尺寸
        - L列(11): 颜色代码
        - M列(12): 颜色名称
        - N列(13): 色号
        - O列(14): 数量
        - Q列(16): 包装数量
        - R列(17): 颜色比例/箱
        """
        ext = os.path.splitext(file_path)[1].lower()
        if ext in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path, header=None)
        elif ext == '.csv':
            df = pd.read_csv(file_path, encoding='utf-8-sig', header=None)
        else:
            raise Exception(f"不支持的文件格式: {ext}")
        
        parsed_data = []
        current_article = None
        
        for idx, row in df.iterrows():
            if len(row) < 18:
                continue
            
            article_col = row.iloc[2]
            
            if pd.notna(article_col) and str(article_col).strip().startswith('CCW'):
                current_article = str(article_col).strip()
                
                parsed_data.append({
                    'row_type': 'main',
                    'article': current_article,
                    'customer_sku': row.iloc[1] if pd.notna(row.iloc[1]) else None,
                    'description': row.iloc[4] if pd.notna(row.iloc[4]) else None,
                    'size': row.iloc[9] if pd.notna(row.iloc[9]) else None,
                    'color_code': row.iloc[11] if pd.notna(row.iloc[11]) else None,
                    'color_name': row.iloc[12] if pd.notna(row.iloc[12]) else None,
                    'color_number': row.iloc[13] if pd.notna(row.iloc[13]) else None,
                    'quantity': row.iloc[14] if pd.notna(row.iloc[14]) else None,  # O列(14) 总数量 -> QTY/SET
                    'box_count': row.iloc[16] if pd.notna(row.iloc[16]) else None,  # Q列(16) 包装数量 -> 件数
                    'box_pcs': row.iloc[20] if pd.notna(row.iloc[20]) else None,  # U列(20) 装箱量 -> PCS每箱数量
                    'color_ratio': row.iloc[17] if pd.notna(row.iloc[17]) else None,
                    'unit_price': row.iloc[22] if pd.notna(row.iloc[22]) else None,  # W列(22) 单价
                    'total_price': row.iloc[23] if pd.notna(row.iloc[23]) else None,  # X列(23) 总价
                    'original_row': idx
                })
            elif pd.notna(row.iloc[11]) and current_article:
                parsed_data.append({
                    'row_type': 'color_detail',
                    'article': current_article,
                    'customer_sku': None,
                    'description': None,
                    'size': None,
                    'color_code': row.iloc[11] if pd.notna(row.iloc[11]) else None,
                    'color_name': row.iloc[12] if pd.notna(row.iloc[12]) else None,
                    'color_number': row.iloc[13] if pd.notna(row.iloc[13]) else None,
                    'quantity': row.iloc[14] if pd.notna(row.iloc[14]) else None,
                    'box_count': None,
                    'box_pcs': None,
                    'color_ratio': row.iloc[17] if pd.notna(row.iloc[17]) else None,
                    'unit_price': None,
                    'total_price': None,
                    'original_row': idx
                })
        
        return parsed_data

    def read_contract(self, file_path):
        """读取销售合同文件"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext in ['.xlsx', '.xls']:
            return pd.read_excel(file_path)
        elif ext == '.csv':
            return pd.read_csv(file_path, encoding='utf-8-sig')
        else:
            raise Exception(f"不支持的文件格式: {ext}")

    def extract_article_numbers(self, data):
        """从数据中提取货号"""
        if isinstance(data, list):
            return list(set([item['article'] for item in data if item.get('article')]))
        
        article_numbers = []
        potential_columns = ['货号', '商品编号', 'SKU', 'Article', 'article', 'item_code', 'Item Code']
        
        for col in potential_columns:
            if col in data.columns:
                articles = data[col].dropna().astype(str).unique().tolist()
                article_numbers.extend([a.strip() for a in articles if a.strip()])
                break
        
        if not article_numbers:
            for col in data.columns:
                for val in data[col].dropna().unique():
                    val_str = str(val).strip()
                    if val_str and len(val_str) > 0:
                        article_numbers.append(val_str)
            article_numbers = list(set(article_numbers))
        
        return article_numbers

    def _process_image(self, image_data, width=None, height=None):
        """处理图片：调整尺寸（如果不指定尺寸则保持原图）"""
        if not image_data:
            return None
        try:
            pil_img = Image.open(io.BytesIO(image_data))
            if pil_img.mode not in ('RGB', 'RGBA'):
                pil_img = pil_img.convert('RGB')
            if width and height:
                pil_img = pil_img.resize((width, height), Image.LANCZOS)
            buffer = io.BytesIO()
            pil_img.save(buffer, format='PNG')
            buffer.seek(0)
            return buffer
        except Exception:
            return None

    def get_images_from_excel(self, file_path, column=None):
        """从Excel文件中提取所有图片，返回 {行号: 图片二进制数据}
        column: 指定列号 (1-indexed)，如不指定则返回所有图片
        """
        images = {}
        try:
            wb = load_workbook(file_path, keep_vba=True, data_only=True)
            ws = wb.active
            
            for img in ws._images:
                try:
                    anchor = img.anchor
                    row = anchor._from.row + 1
                    col = anchor._from.col + 1
                    
                    # 如果指定了列，只返回该列的图片
                    if column is not None and col != column:
                        continue
                    
                    # _data is a method that returns bytes
                    try:
                        img_data = img._data()
                        if img_data:
                            images[row] = img_data
                    except Exception:
                        pass
                except Exception:
                    pass
            
            wb.close()
            return images
        except Exception:
            return images

    def _extract_image_from_excel(self, file_path, row_index, column_index=4):
        """从Excel文件中提取图片（二进制数据）"""
        images = self.get_images_from_excel(file_path)
        return images.get(row_index)

    def _create_duizhangdan_sheet(self, wb, all_data, selected_articles):
        """创建对账单工作表"""
        ws = wb.create_sheet("对账单")
        
        # 冻结窗格在第3行
        ws.freeze_panes = 'A3'
        
        # 设置视图为分页预览（显示灰色周边）
        ws.sheet_view.view = 'pageBreakPreview'
        
        # 隐藏网格线
        ws.sheet_view.showGridLines = False
        
        selected_set = set(str(a) for a in selected_articles)
        
        title_font = Font(name='Trebuchet MS', bold=True, size=12)
        header_font = Font(name='Trebuchet MS', size=12)
        data_font = Font(name='Trebuchet MS', size=12)
        article_font = Font(name='Times New Roman', size=14)
        header_alignment = Alignment(horizontal='center', vertical='center')
        data_alignment = Alignment(horizontal='center', vertical='center')
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Row 1: 标题
        ws.merge_cells('D1:G1')
        ws['D1'] = 'LISA 10月30日对账单'
        ws['D1'].font = title_font
        ws['D1'].alignment = header_alignment
        
        ws.merge_cells('H1:H1')
        ws['H1'] = '包装:纸箱+隔板'
        ws['H1'].font = title_font
        ws['H1'].alignment = header_alignment
        
        ws.merge_cells('I1:I1')
        ws['I1'] = '日期'
        ws['I1'].font = title_font
        ws['I1'].alignment = header_alignment
        
        ws.row_dimensions[1].height = 42
        
        # Row 2: 列标题 - 绿色背景
        green_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF92D050'))
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        headers = ['合同号', '客户名称', '客户货号', '图片', '件数', '包装数', '总数', '单价', '总价']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = green_fill
            cell.border = thin_border
        
        ws.row_dimensions[2].height = 27
        
        # 设置列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        
        # 数据行
        row_num = 3
        current_contract = None
        current_customer = None
        contract_start_row = None
        customer_start_row = None
        
        # 按货号分组，只取主货号行
        articles_seen = set()
        
        for contract_info in all_data:
            file_path = contract_info['file_path']
            rows = contract_info['rows']
            images = contract_info.get('images', {})
            
            customer_name, contract_no, _ = self.get_contract_info(file_path)
            
            for item in rows:
                article = item.get('article')
                if article not in selected_set:
                    continue
                
                # 只处理主货号行，避免重复
                if article in articles_seen:
                    continue
                articles_seen.add(article)
                
                article_display = article
                
                # 件数 = 总箱数, 包装数 = PCS每箱数量
                box_count = item.get('box_count', 0) or 0
                box_pcs = item.get('box_pcs', 0) or 0
                unit_price = item.get('unit_price', 0) or 0
                
                # 检查是否需要合并A列（同一合同的货号）
                if current_contract != contract_no:
                    # 合并上一组的A列
                    if current_contract is not None and contract_start_row is not None:
                        ws.merge_cells(f'A{contract_start_row}:A{row_num-1}')
                    contract_start_row = row_num
                    current_contract = contract_no
                
                # 检查是否需要合并B列（同一客户的货号）
                if current_customer != customer_name:
                    # 合并上一组的B列
                    if current_customer is not None and customer_start_row is not None:
                        ws.merge_cells(f'B{customer_start_row}:B{row_num-1}')
                    customer_start_row = row_num
                    current_customer = customer_name
                
                # A列: 合同号
                cell = ws.cell(row=row_num, column=1, value=contract_no)
                cell.font = data_font
                cell.alignment = data_alignment
                
                # B列: 客户名称
                cell = ws.cell(row=row_num, column=2, value=customer_name)
                cell.font = data_font
                cell.alignment = data_alignment
                
                # C列: 客户货号 - 使用Times New Roman, Size 14
                cell = ws.cell(row=row_num, column=3, value=article_display)
                cell.font = article_font
                cell.alignment = data_alignment
                
                # E列: 件数
                cell = ws.cell(row=row_num, column=5, value=box_count)
                cell.font = data_font
                cell.alignment = data_alignment
                
                # F列: 包装数
                cell = ws.cell(row=row_num, column=6, value=box_pcs)
                cell.font = data_font
                cell.alignment = data_alignment
                
                # G列: 总数 = 件数 * 包装数
                cell = ws.cell(row=row_num, column=7, value=f'=E{row_num}*F{row_num}')
                cell.font = data_font
                cell.alignment = data_alignment
                
                # H列: 单价
                cell = ws.cell(row=row_num, column=8, value=unit_price)
                cell.font = data_font
                cell.alignment = data_alignment
                
                # I列: 总价 = 总数 * 单价
                cell = ws.cell(row=row_num, column=9, value=f'=G{row_num}*H{row_num}')
                cell.font = data_font
                cell.alignment = data_alignment
                
                # 添加图片 - 使用原图，设置显示大小为80x80
                excel_row = item.get('original_row', 0) + 1
                if excel_row in images:
                    img_data = images[excel_row]
                    if img_data:
                        buffer = io.BytesIO(img_data)
                        img = XLImage(buffer)
                        img.width = 80
                        img.height = 80
                        ws.add_image(img, f'D{row_num}')
                        # 设置行高以适应图片
                        ws.row_dimensions[row_num].height = 92.1
                else:
                    ws.row_dimensions[row_num].height = 92.1
                
                row_num += 1
        
        # 合并最后一组
        if contract_start_row is not None:
            ws.merge_cells(f'A{contract_start_row}:A{row_num-1}')
        if customer_start_row is not None:
            ws.merge_cells(f'B{customer_start_row}:B{row_num-1}')
        
        # 为数据行添加边框
        for row in range(3, row_num):
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = data_border
        
        # Row N: 合计
        total_row = row_num
        total_font = Font(name='Trebuchet MS', bold=True, size=12)
        total_alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells(f'A{total_row}:D{total_row}')
        cell = ws[f'A{total_row}']
        cell.value = '合计'
        cell.font = total_font
        cell.alignment = total_alignment
        cell.border = data_border
        
        cell = ws[f'E{total_row}']
        cell.value = f'=SUM(E3:E{total_row-1})'
        cell.font = total_font
        cell.alignment = total_alignment
        cell.border = data_border
        
        cell = ws[f'G{total_row}']
        cell.value = f'=SUM(G3:G{total_row-1})'
        cell.font = total_font
        cell.alignment = total_alignment
        cell.border = data_border
        
        cell = ws[f'I{total_row}']
        cell.value = f'=SUM(I3:I{total_row-1})'
        cell.font = total_font
        cell.alignment = total_alignment
        cell.border = data_border
        
        ws.row_dimensions[total_row].height = 32.1
        
        # Row N+1: 应收款
        receivable_row = total_row + 1
        receivable_font = Font(name='Trebuchet MS', bold=True, size=12)
        
        ws.merge_cells(f'A{receivable_row}:H{receivable_row}')
        cell = ws[f'A{receivable_row}']
        cell.value = '应收款:'
        cell.font = receivable_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = data_border
        
        cell = ws[f'I{receivable_row}']
        cell.value = f'=I{total_row}'
        cell.font = receivable_font
        cell.alignment = total_alignment
        
        ws.row_dimensions[receivable_row].height = 32.1

    def validate_contract(self, file_path):
        """验证合同文件"""
        if not os.path.exists(file_path):
            return False, "文件不存在"
        
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in ['.xlsx', '.xls', '.csv']:
            return False, "请选择 Excel 或 CSV 文件"
        
        return True, ""

    def get_contract_info(self, file_path):
        """从销售合同中提取客户名称和PI号"""
        try:
            ext = os.path.splitext(file_path)[1].lower()
            if ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path, header=None)
            elif ext == '.csv':
                df = pd.read_csv(file_path, encoding='utf-8-sig', header=None)
            else:
                return None, None, None
            
            # 客户名称：第4行C列(2)
            customer_name = df.iloc[4, 2] if len(df.columns) > 2 and pd.notna(df.iloc[4, 2]) else None
            if customer_name:
                customer_name = str(customer_name).strip()
            
            # 合同号：第4行W列(22)
            contract_no = df.iloc[4, 22] if len(df.columns) > 22 and pd.notna(df.iloc[4, 22]) else None
            if contract_no:
                contract_no = str(contract_no).strip()
            else:
                # 如果找不到，从文件名提取
                contract_no = os.path.basename(file_path)
                contract_no = contract_no.replace('#CCW#PI_', '').replace('#CCW#', '')
                contract_no = os.path.splitext(contract_no)[0]
            
            return customer_name, contract_no, contract_no
        except:
            return None, None, None
