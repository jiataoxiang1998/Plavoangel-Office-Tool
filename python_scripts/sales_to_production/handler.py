import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color, Border, Side


class SalesToProductionHandler:
    def __init__(self):
        self.sales_contracts = {}
        self.selected_articles = set()

    def add_contract(self, file_path):
        try:
            parsed_data = self.read_sales_contract(file_path)
            article_numbers = []
            seen = set()
            for item in parsed_data:
                art = item.get('article')
                if art and art not in seen:
                    seen.add(art)
                    article_numbers.append(art)
            
            self.sales_contracts[file_path] = {
                'article_numbers': article_numbers,
                'data': parsed_data
            }
            
            for art in article_numbers:
                self.selected_articles.add(art)
            
            return True, article_numbers
        except Exception as e:
            return False, str(e)

    def remove_contract(self, file_path):
        if file_path in self.sales_contracts:
            removed_articles = self.sales_contracts[file_path]['article_numbers']
            del self.sales_contracts[file_path]
            self.selected_articles = self.selected_articles - set(removed_articles)
            return True
        return False

    def get_all_article_numbers(self):
        all_articles = []
        seen = set()
        for contract in self.sales_contracts.values():
            for art in contract['article_numbers']:
                if art not in seen:
                    seen.add(art)
                    all_articles.append(art)
        return all_articles

    def toggle_article_selection(self, article_number):
        if article_number in self.selected_articles:
            self.selected_articles.discard(article_number)
            return False
        else:
            self.selected_articles.add(article_number)
            return True

    def get_selected_articles(self):
        return list(self.selected_articles)

    def is_article_selected(self, article_number):
        return article_number in self.selected_articles

    def get_selected_articles_in_order(self):
        all_articles = self.get_all_article_numbers()
        return [art for art in all_articles if art in self.selected_articles]

    def generate_production_order(self, output_path, selected_articles=None):
        if selected_articles is None:
            selected_articles = self.get_selected_articles_in_order()
        
        if not selected_articles:
            raise Exception("请先选择货号")
        
        all_data = []
        selected_set = set(str(a) for a in selected_articles)
        
        for contract_path, contract_data in self.sales_contracts.items():
            parsed_data = contract_data['data']
            matching_rows = []
            
            if isinstance(parsed_data, list):
                for item in parsed_data:
                    if item.get('article') in selected_set:
                        matching_rows.append(item)
            
            if matching_rows:
                all_data.append({
                    'file_path': contract_path,
                    'rows': matching_rows
                })
        
        if not all_data:
            raise Exception("未找到选中的货号数据")
        
        base_name, ext = os.path.splitext(output_path)
        if not ext:
            ext = '.xlsx'
            output_path = base_name + ext
        
        wb = Workbook()
        self._create_production_sheet(wb, all_data, selected_articles)
        
        wb.save(output_path)
        return output_path

    def read_sales_contract(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        if ext in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path, header=None)
        elif ext == '.csv':
            df = pd.read_csv(file_path, encoding='utf-8-sig', header=None)
        else:
            raise Exception(f"不支持的文件格式: {ext}")
        
        parsed_data = []
        current_article = None
        
        for idx, row in df.iterrows():
            if len(row) < 19:
                continue
            
            article_col = row.iloc[2]  # Column C
            
            if pd.notna(article_col) and str(article_col).strip().startswith('CCW'):
                current_article = str(article_col).strip()
                
                # K: JK8622&F779 -> split to get supplier info
                supplier_code = ''
                supplier_material = ''
                if pd.notna(row.iloc[10]):
                    supplier_parts = str(row.iloc[10]).split('&')
                    supplier_code = supplier_parts[0].strip() if supplier_parts else ''
                    supplier_material = supplier_parts[1].strip() if len(supplier_parts) > 1 else ''
                
                parsed_data.append({
                    'row_type': 'main',
                    'article': current_article,
                    'customer_sku': row.iloc[3] if pd.notna(row.iloc[3]) else None,  # D: 条形码
                    'size': row.iloc[9] if pd.notna(row.iloc[9]) else None,  # J: 尺寸
                    'supplier_code': supplier_code,  # I: 供应商编码
                    'supplier_material': supplier_material,  # L: 供应商物料编号 (F779 from K&F779)
                    'color_code': row.iloc[12] if pd.notna(row.iloc[12]) else None,  # M: 色号
                    'color_name': row.iloc[12] if pd.notna(row.iloc[12]) else None,  # M: 色号
                    'material_main': row.iloc[13] if pd.notna(row.iloc[13]) else None,  # N: 双惠主料
                    'material_sub': row.iloc[14] if pd.notna(row.iloc[14]) else None,  # O: 双惠配料
                    'color_ratio': row.iloc[18] if pd.notna(row.iloc[18]) else None,  # S: 中盒比例
                    'quantity': row.iloc[15] if pd.notna(row.iloc[15]) else None,  # P: 数量
                    'original_row': idx
                })
            elif pd.notna(row.iloc[11]) and current_article:  # L列有颜色代码
                parsed_data.append({
                    'row_type': 'color_detail',
                    'article': current_article,
                    'color_code': row.iloc[12] if pd.notna(row.iloc[12]) else None,  # M: 色号
                    'color_name': row.iloc[12] if pd.notna(row.iloc[12]) else None,  # M: 色号
                    'material_main': row.iloc[13] if pd.notna(row.iloc[13]) else None,  # N: 双惠主料
                    'material_sub': row.iloc[14] if pd.notna(row.iloc[14]) else None,  # O: 双惠配料
                    'color_ratio': row.iloc[18] if pd.notna(row.iloc[18]) else None,  # S: 中盒比例
                    'quantity': row.iloc[15] if pd.notna(row.iloc[15]) else None,  # P: 数量
                    'original_row': idx
                })
        
        return parsed_data

    def _create_production_sheet(self, wb, all_data, selected_articles):
        ws = wb.active
        ws.title = "生产单"
        
        ws.sheet_view.showGridLines = False
        
        thin_font = Font(name='Trebuchet MS', size=9)
        header_font = Font(name='Trebuchet MS', size=9, bold=True)
        
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF92D050'))
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25
        
        ws.cell(1, 1).value = '客户货号'
        ws.cell(1, 2).value = '条形码'
        ws.cell(1, 3).value = '内阁编号 & 尺寸'
        ws.cell(1, 4).value = '颜色'
        ws.cell(1, 5).value = ''
        ws.cell(1, 6).value = '数量'
        ws.cell(1, 7).value = '中盒比例'
        ws.cell(1, 8).value = '包身后幅+前幅'
        ws.cell(1, 9).value = ''
        ws.cell(1, 10).value = ''
        ws.cell(1, 11).value = '大面拼接块'
        ws.cell(1, 12).value = ''
        ws.cell(1, 13).value = ''
        
        for col in range(1, 14):
            ws.cell(1, col).font = header_font
            ws.cell(1, col).alignment = header_alignment
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).border = thin_border
        
        ws.cell(2, 4).value = '客称'
        ws.cell(2, 5).value = '中'
        
        ws.cell(2, 8).value = '供应商'
        ws.cell(2, 9).value = '材料'
        ws.cell(2, 10).value = '色号'
        
        ws.cell(2, 11).value = '供应商'
        ws.cell(2, 12).value = '材料'
        ws.cell(2, 13).value = '色号'
        
        for col in [4, 5, 8, 9, 10, 11, 12, 13]:
            ws.cell(2, col).font = header_font
            ws.cell(2, col).alignment = header_alignment
            ws.cell(2, col).fill = header_fill
            ws.cell(2, col).border = thin_border
        
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:B2')
        ws.merge_cells('C1:C2')
        ws.merge_cells('D1:E1')
        ws.merge_cells('F1:F2')
        ws.merge_cells('G1:G2')
        ws.merge_cells('H1:J1')
        ws.merge_cells('K1:M1')
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 6
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 14
        ws.column_dimensions['M'].width = 14
        
        row_num = 3
        
        selected_set = set(str(a) for a in selected_articles)
        
        articles_data = {}
        for contract_info in all_data:
            for item in contract_info['rows']:
                article = item.get('article')
                if article not in selected_set:
                    continue
                if article not in articles_data:
                    articles_data[article] = {'items': []}
                articles_data[article]['items'].append(item)
        
        for article in selected_articles:
            if article not in articles_data:
                continue
            
            items = articles_data[article]['items']
            
            main_item = None
            color_items = []
            for item in items:
                if item.get('row_type') == 'main':
                    main_item = item
                else:
                    color_items.append(item)
            
            if not main_item:
                continue
            
            quantity = main_item.get('quantity', 0) or 0
            
            article_start_row = row_num
            
            if color_items:
                main_color_code = str(main_item.get('color_code', '')) if main_item.get('color_code') else ''
                main_color_name = str(main_item.get('color_name', '')) if main_item.get('color_name') else ''
                supplier_code = main_item.get('supplier_code', '')
                supplier_material = main_item.get('supplier_material', '')
                color_ratio = main_item.get('color_ratio', 0) or 0
                material_main = str(main_item.get('material_main', '')) if main_item.get('material_main') else ''
                material_sub = str(main_item.get('material_sub', '')) if main_item.get('material_sub') else ''
                
                ws.cell(row_num, 1, article).font = thin_font
                ws.cell(row_num, 1).alignment = data_alignment
                ws.cell(row_num, 2, main_item.get('customer_sku', '')).font = thin_font
                ws.cell(row_num, 2).alignment = data_alignment
                ws.cell(row_num, 3, str(main_item.get('size', ''))).font = thin_font
                ws.cell(row_num, 3).alignment = data_alignment
                
                ws.cell(row_num, 4, main_color_code).font = thin_font
                ws.cell(row_num, 4).alignment = data_alignment
                ws.cell(row_num, 5, main_color_name).font = thin_font
                ws.cell(row_num, 5).alignment = data_alignment
                ws.cell(row_num, 6, quantity).font = thin_font
                ws.cell(row_num, 6).alignment = data_alignment
                ws.cell(row_num, 7, color_ratio).font = thin_font
                ws.cell(row_num, 7).alignment = data_alignment
                
                ws.cell(row_num, 8, '').font = thin_font
                ws.cell(row_num, 8).alignment = data_alignment
                ws.cell(row_num, 9, material_main).font = thin_font
                ws.cell(row_num, 9).alignment = data_alignment
                ws.cell(row_num, 10, main_color_code).font = thin_font
                ws.cell(row_num, 10).alignment = data_alignment
                
                ws.cell(row_num, 11, '').font = thin_font
                ws.cell(row_num, 11).alignment = data_alignment
                ws.cell(row_num, 12, material_sub).font = thin_font
                ws.cell(row_num, 12).alignment = data_alignment
                ws.cell(row_num, 13, main_color_code).font = thin_font
                ws.cell(row_num, 13).alignment = data_alignment
                
                for col in range(1, 14):
                    ws.cell(row_num, col).border = thin_border
                
                row_num += 1
                
                for idx, color_item in enumerate(color_items):
                    color_code = str(color_item.get('color_code', '')) if color_item.get('color_code') else ''
                    color_name = str(color_item.get('color_name', '')) if color_item.get('color_name') else ''
                    color_qty = color_item.get('quantity', 0) or 0
                    material_main_c = str(color_item.get('material_main', '')) if color_item.get('material_main') else ''
                    material_sub_c = str(color_item.get('material_sub', '')) if color_item.get('material_sub') else ''
                    color_ratio_c = color_item.get('color_ratio', 0) or 0
                    if not color_ratio_c and main_item.get('color_ratio'):
                        color_ratio_c = main_item.get('color_ratio')
                    
                    ws.cell(row_num, 1, '').font = thin_font
                    ws.cell(row_num, 1).alignment = data_alignment
                    ws.cell(row_num, 2, '').font = thin_font
                    ws.cell(row_num, 2).alignment = data_alignment
                    ws.cell(row_num, 3, '').font = thin_font
                    ws.cell(row_num, 3).alignment = data_alignment
                    
                    ws.cell(row_num, 4, color_code).font = thin_font
                    ws.cell(row_num, 4).alignment = data_alignment
                    ws.cell(row_num, 5, color_name).font = thin_font
                    ws.cell(row_num, 5).alignment = data_alignment
                    ws.cell(row_num, 6, color_qty).font = thin_font
                    ws.cell(row_num, 6).alignment = data_alignment
                    ws.cell(row_num, 7, color_ratio_c).font = thin_font
                    ws.cell(row_num, 7).alignment = data_alignment
                    
                    ws.cell(row_num, 8, '').font = thin_font
                    ws.cell(row_num, 8).alignment = data_alignment
                    ws.cell(row_num, 9, material_main_c).font = thin_font
                    ws.cell(row_num, 9).alignment = data_alignment
                    ws.cell(row_num, 10, color_code).font = thin_font
                    ws.cell(row_num, 10).alignment = data_alignment
                    
                    ws.cell(row_num, 11, '').font = thin_font
                    ws.cell(row_num, 11).alignment = data_alignment
                    ws.cell(row_num, 12, material_sub_c).font = thin_font
                    ws.cell(row_num, 12).alignment = data_alignment
                    ws.cell(row_num, 13, color_code).font = thin_font
                    ws.cell(row_num, 13).alignment = data_alignment
                    
                    for col in range(1, 14):
                        ws.cell(row_num, col).border = thin_border
                    
                    row_num += 1
            else:
                color_code = str(main_item.get('color_code', '')) if main_item.get('color_code') else ''
                color_name = str(main_item.get('color_name', '')) if main_item.get('color_name') else ''
                material_main = str(main_item.get('material_main', '')) if main_item.get('material_main') else ''
                material_sub = str(main_item.get('material_sub', '')) if main_item.get('material_sub') else ''
                supplier_code = main_item.get('supplier_code', '')
                supplier_material = main_item.get('supplier_material', '')
                color_ratio = main_item.get('color_ratio', 0) or 0
                
                ws.cell(row_num, 1, article).font = thin_font
                ws.cell(row_num, 1).alignment = data_alignment
                ws.cell(row_num, 2, main_item.get('customer_sku', '')).font = thin_font
                ws.cell(row_num, 2).alignment = data_alignment
                ws.cell(row_num, 3, str(main_item.get('size', ''))).font = thin_font
                ws.cell(row_num, 3).alignment = data_alignment
                
                ws.cell(row_num, 4, color_code).font = thin_font
                ws.cell(row_num, 4).alignment = data_alignment
                ws.cell(row_num, 5, color_name).font = thin_font
                ws.cell(row_num, 5).alignment = data_alignment
                ws.cell(row_num, 6, quantity).font = thin_font
                ws.cell(row_num, 6).alignment = data_alignment
                ws.cell(row_num, 7, color_ratio).font = thin_font
                ws.cell(row_num, 7).alignment = data_alignment
                
                ws.cell(row_num, 8, '').font = thin_font
                ws.cell(row_num, 8).alignment = data_alignment
                ws.cell(row_num, 9, material_main).font = thin_font
                ws.cell(row_num, 9).alignment = data_alignment
                ws.cell(row_num, 10, color_code).font = thin_font
                ws.cell(row_num, 10).alignment = data_alignment
                
                ws.cell(row_num, 11, '').font = thin_font
                ws.cell(row_num, 11).alignment = data_alignment
                ws.cell(row_num, 12, material_sub).font = thin_font
                ws.cell(row_num, 12).alignment = data_alignment
                ws.cell(row_num, 13, color_code).font = thin_font
                ws.cell(row_num, 13).alignment = data_alignment
                
                for col in range(1, 14):
                    ws.cell(row_num, col).border = thin_border
                
                row_num += 1
            
            article_end_row = row_num - 1
            if article_end_row > article_start_row:
                ws.merge_cells(f'A{article_start_row}:A{article_end_row}')
                ws.merge_cells(f'B{article_start_row}:B{article_end_row}')
                ws.merge_cells(f'C{article_start_row}:C{article_end_row}')
            
            row_num += 1

    def validate_contract(self, file_path):
        if not os.path.exists(file_path):
            return False, "文件不存在"
        
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in ['.xlsx', '.xls', '.csv']:
            return False, "请选择 Excel 或 CSV 文件"
        
        return True, ""

    def get_contract_info(self, file_path):
        try:
            ext = os.path.splitext(file_path)[1].lower()
            if ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path, header=None)
            elif ext == '.csv':
                df = pd.read_csv(file_path, encoding='utf-8-sig', header=None)
            else:
                return None, None, None
            
            customer_name = df.iloc[4, 2] if len(df.columns) > 2 and pd.notna(df.iloc[4, 2]) else None
            if customer_name:
                customer_name = str(customer_name).strip()
            
            contract_no = df.iloc[4, 22] if len(df.columns) > 22 and pd.notna(df.iloc[4, 22]) else None
            if contract_no:
                contract_no = str(contract_no).strip()
            else:
                contract_no = os.path.basename(file_path)
                contract_no = contract_no.replace('#CCW#PI_', '').replace('#CCW#', '')
                contract_no = os.path.splitext(contract_no)[0]
            
            return customer_name, contract_no, contract_no
        except:
            return None, None, None